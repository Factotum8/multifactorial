# -*- encoding: utf-8 -*-
import math as m

from openpyxl import load_workbook

table_G = 0.5157


def main():
    wb = load_workbook('./data_matrix.xlsx')
    pm = wb.get_sheet_by_name(wb.get_sheet_names()[0]) # planning matrix
    # print (pm.title)
    print ("    Матрица планирования")
    X1 = []; X2 = []; X3 = []
    Y1 = []; Y2 = []; Y3 = []; Yn = []
    for i in range(3, 11):
        print(pm.cell(row=i, column=1).value, pm.cell(row=i, column=3).value,
              pm.cell(row=i, column=4).value, pm.cell(row=i, column=5).value,
              "%.3f" % pm.cell(row=i, column=19).value,
              "%.3f" % pm.cell(row=i, column=20).value,
              "%.3f" % pm.cell(row=i, column=21).value,
              "%.3f" % pm.cell(row=i, column=22).value)

        X1.append(pm.cell(row=i, column=3).value)
        X2.append(pm.cell(row=i, column=4).value)
        X3.append(pm.cell(row=i, column=5).value)

        Y1.append(pm.cell(row=i, column=19).value)
        Y2.append(pm.cell(row=i, column=20).value)
        Y3.append(pm.cell(row=i, column=21).value)
        Yn.append(pm.cell(row=i, column=22).value)

    print("\n    Вычесления")
    S2j = [] # среднее значени параметра оптимизации
    Sj = [] # cреднеквадратичное отклонение ??! или это дисперсия
    for y1, y2, y3, yn in zip(Y1, Y2, Y3, Yn):
        current_element = 0.5 * (m.pow((y1 - yn), 2) + m.pow((y2 - yn), 2) + m.pow((y3 - yn), 2))
        s = m.sqrt(current_element)
        S2j.append(current_element)
        Sj.append(s)
        print("S^2 = %.3f " % current_element, "   S = %.3f " % s)

    # Критерий Кохнера
    G = (max(S2j))/1
    if G < table_G:
        print("\n    Дисперсия однородна G = %.3f" % G)
    else:
        print("    Дисперсия не однородна G = %.3f" % G)

    SY = len(S2j) * sum(S2j) # Остаточная дисперсия
    print("    Остаточная дисперсия SY =  %.3f" % SY)



if __name__ == '__main__':
    main()